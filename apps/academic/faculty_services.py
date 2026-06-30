"""
Faculty Master & Teaching Allocation Import Services.

Two-phase import pattern (validate → process_import) consistent with
the existing AcademicStructureImportService.

FacultyMasterImportService:
    Parses staff list Excel/CSV, detects columns dynamically,
    upserts FacultyMaster records by email within the Academic Year,
    and auto-links to User accounts.

TeachingAllocationImportService:
    Parses course allocation Excel/CSV with multi-row headers,
    resolves faculty aliases via FacultyMaster short_form or name,
    matches Course Code → SemesterSubject, and creates
    FacultyTeachingAssignment records with theory vs practical-batch
    distinction.
"""
import csv
import io
import re
from django.db import transaction
from django.contrib.auth import get_user_model

from apps.academic.models import (
    AcademicYear, Semester, SemesterSubject,
    FacultyMaster, FacultyTeachingAssignment, FacultyImportLog,
)

User = get_user_model()


# ═══════════════════════════════════════════════════════
# FACULTY MASTER IMPORT
# ═══════════════════════════════════════════════════════

class FacultyMasterImportService:
    """
    Parses an uploaded Excel/CSV staff list and creates/updates
    FacultyMaster records for the given Academic Year.

    Supports flexible column detection — matches common header names
    instead of requiring exact column names.
    """

    # Patterns to match column headers (case-insensitive)
    NAME_PATTERNS = ['name of faculty', 'faculty name', 'name', 'full name', 'staff name']
    ALIAS_PATTERNS = ['alias', 'short form', 'short name', 'abbreviation', 'initials', 'code']
    EMAIL_PATTERNS = ['email id', 'email', 'official email', 'mail', 'e-mail']
    EMPCODE_PATTERNS = ['employee code', 'emp code', 'emp id', 'employee id', 'staff id']
    DEPT_PATTERNS = ['department', 'dept', 'dept.']

    def __init__(self, file_obj, filename, academic_year):
        self.file_obj = file_obj
        self.filename = filename
        self.academic_year = academic_year
        self.errors = []
        self.warnings = []
        self.valid_rows = []
        self.file_format = 'xlsx' if filename.lower().endswith('.xlsx') else 'csv'

        # Pre-load user cache for auto-linking
        self.users_cache = {
            u.email.lower(): u
            for u in User.objects.exclude(email='')
        }

    def validate(self):
        """Phase 1: Parse and validate without writing to DB."""
        self.errors = []
        self.warnings = []
        self.valid_rows = []

        try:
            if self.file_format == 'xlsx':
                self._parse_xlsx()
            else:
                self._parse_csv()
        except Exception as e:
            self.errors.append({
                'row': 0,
                'message': f"Failed to parse file: {str(e)}"
            })

        return self._get_summary()

    def _parse_xlsx(self):
        """Parse Excel (.xlsx) staff list."""
        import openpyxl

        self.file_obj.seek(0)
        wb = openpyxl.load_workbook(self.file_obj, read_only=True, data_only=True)

        # Try to find the TEACHING sheet, fall back to active
        ws = None
        for sheet_name in wb.sheetnames:
            if 'teaching' in sheet_name.lower():
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.errors.append({'row': 0, 'message': "File is empty."})
            return

        # Find header row
        header_idx, col_map = self._find_header_row(rows)
        if header_idx is None:
            self.errors.append({
                'row': 0,
                'message': (
                    "Could not find header row. Expected columns like: "
                    "Name of Faculty, Alias, Email ID."
                )
            })
            return

        if 'name' not in col_map:
            self.errors.append({
                'row': 0,
                'message': "Could not find a 'Faculty Name' column."
            })
            return

        if 'email' not in col_map:
            self.errors.append({
                'row': 0,
                'message': "Could not find an 'Email ID' column."
            })
            return

        # Parse data rows
        seen_emails = set()
        seen_aliases = set()

        for row_idx in range(header_idx + 1, len(rows)):
            row = rows[row_idx]
            excel_row = row_idx + 1  # 1-indexed for user-facing messages

            # Skip completely empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            name_val = self._get_cell(row, col_map.get('name'))
            alias_val = self._get_cell(row, col_map.get('alias'))
            email_val = self._get_cell(row, col_map.get('email'))
            empcode_val = self._get_cell(row, col_map.get('empcode'))
            dept_val = self._get_cell(row, col_map.get('dept'))

            # Skip rows without a name (likely separator/summary)
            if not name_val:
                continue

            # Validate email
            if not email_val:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': "Missing email ID. Every faculty must have an email."
                })
                continue

            email_lower = email_val.strip().lower()

            # Basic email format check
            if '@' not in email_lower:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': f"Invalid email format: '{email_val}'"
                })
                continue

            # Duplicate email within this file
            if email_lower in seen_emails:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': f"Duplicate email '{email_val}' in this file."
                })
                continue
            seen_emails.add(email_lower)

            # Alias duplicate check (only if alias is present)
            alias_upper = alias_val.strip().upper() if alias_val else ''
            if alias_upper:
                if alias_upper in seen_aliases:
                    self.warnings.append({
                        'row': excel_row,
                        'identifier': name_val,
                        'message': f"Duplicate alias '{alias_val}' in this file."
                    })
                seen_aliases.add(alias_upper)

            self.valid_rows.append({
                'row': excel_row,
                'faculty_name': name_val.strip(),
                'short_form': alias_val.strip() if alias_val else '',
                'email': email_lower,
                'employee_code': empcode_val.strip() if empcode_val else '',
                'department': dept_val.strip() if dept_val else '',
            })

        wb.close()

    def _parse_csv(self):
        """Parse CSV staff list."""
        self.file_obj.seek(0)
        try:
            decoded = self.file_obj.read().decode('utf-8', errors='replace')
        except AttributeError:
            decoded = self.file_obj.read()

        io_string = io.StringIO(decoded)
        reader = csv.reader(io_string)
        all_rows = list(reader)

        if not all_rows:
            self.errors.append({'row': 0, 'message': "File is empty."})
            return

        # Convert to tuples for uniform handling
        rows_as_tuples = [tuple(row) for row in all_rows]
        header_idx, col_map = self._find_header_row(rows_as_tuples)

        if header_idx is None or 'name' not in col_map or 'email' not in col_map:
            self.errors.append({
                'row': 0,
                'message': "Could not find header row with Faculty Name and Email columns."
            })
            return

        seen_emails = set()
        seen_aliases = set()

        for row_idx in range(header_idx + 1, len(rows_as_tuples)):
            row = rows_as_tuples[row_idx]
            excel_row = row_idx + 1

            if not any(str(c).strip() for c in row):
                continue

            name_val = self._get_cell(row, col_map.get('name'))
            alias_val = self._get_cell(row, col_map.get('alias'))
            email_val = self._get_cell(row, col_map.get('email'))
            empcode_val = self._get_cell(row, col_map.get('empcode'))
            dept_val = self._get_cell(row, col_map.get('dept'))

            if not name_val:
                continue

            if not email_val:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': "Missing email ID."
                })
                continue

            email_lower = email_val.strip().lower()
            if '@' not in email_lower:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': f"Invalid email: '{email_val}'"
                })
                continue

            if email_lower in seen_emails:
                self.errors.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': f"Duplicate email '{email_val}'."
                })
                continue
            seen_emails.add(email_lower)

            alias_upper = alias_val.strip().upper() if alias_val else ''
            if alias_upper and alias_upper in seen_aliases:
                self.warnings.append({
                    'row': excel_row,
                    'identifier': name_val,
                    'message': f"Duplicate alias '{alias_val}'."
                })
            if alias_upper:
                seen_aliases.add(alias_upper)

            self.valid_rows.append({
                'row': excel_row,
                'faculty_name': name_val.strip(),
                'short_form': alias_val.strip() if alias_val else '',
                'email': email_lower,
                'employee_code': empcode_val.strip() if empcode_val else '',
                'department': dept_val.strip() if dept_val else '',
            })

    def _find_header_row(self, rows):
        """
        Scan rows to find the header row by matching known column patterns.
        Returns (row_index, column_map) or (None, {}).
        """
        for idx, row in enumerate(rows[:10]):  # Check first 10 rows
            row_strs = [str(cell).strip().lower() if cell else '' for cell in row]

            col_map = {}
            for col_idx, val in enumerate(row_strs):
                if not val:
                    continue
                if any(p in val for p in self.NAME_PATTERNS):
                    col_map['name'] = col_idx
                elif any(p == val for p in self.ALIAS_PATTERNS):
                    col_map['alias'] = col_idx
                elif any(p in val for p in self.EMAIL_PATTERNS):
                    col_map['email'] = col_idx
                elif any(p in val for p in self.EMPCODE_PATTERNS):
                    col_map['empcode'] = col_idx
                elif any(p in val for p in self.DEPT_PATTERNS):
                    col_map['dept'] = col_idx

            # Consider it a valid header if we found at least name and email
            if 'name' in col_map and 'email' in col_map:
                return idx, col_map

        return None, {}

    def _get_cell(self, row, col_idx):
        """Safely get a cell value."""
        if col_idx is None or col_idx >= len(row):
            return None
        val = row[col_idx]
        if val is None:
            return None
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() == 'none':
            return None
        return val_str

    def _get_summary(self):
        """Build a summary dict."""
        return {
            'total_valid': len(self.valid_rows),
            'total_errors': len(self.errors),
            'total_warnings': len(self.warnings),
            'errors': self.errors[:100],
            'warnings': self.warnings[:50],
            'valid_sample': self.valid_rows[:15],
            'file_format': self.file_format,
        }

    @transaction.atomic
    def process_import(self, imported_by=None):
        """
        Phase 2: Commit validated data to the database.
        Upserts by email — creates new records or updates existing ones.
        """
        if not self.valid_rows:
            return FacultyImportLog.objects.create(
                academic_year=self.academic_year,
                import_type=FacultyImportLog.ImportType.FACULTY_MASTER,
                original_filename=self.filename,
                file_format=self.file_format,
                imported_by=imported_by,
                status=FacultyImportLog.ImportStatus.FAILED,
                summary={'error': 'No valid rows to import.'},
                error_log=self.errors,
            )

        created_count = 0
        updated_count = 0

        for row_data in self.valid_rows:
            email = row_data['email']

            # Auto-link to User account if exists
            linked_user = self.users_cache.get(email)

            defaults = {
                'faculty_name': row_data['faculty_name'],
                'short_form': row_data['short_form'],
                'employee_code': row_data['employee_code'],
                'department': row_data['department'],
                'is_active': True,
                'user': linked_user,
            }

            obj, created = FacultyMaster.objects.update_or_create(
                academic_year=self.academic_year,
                email=email,
                defaults=defaults,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        import_log = FacultyImportLog.objects.create(
            academic_year=self.academic_year,
            import_type=FacultyImportLog.ImportType.FACULTY_MASTER,
            original_filename=self.filename,
            file_format=self.file_format,
            imported_by=imported_by,
            status=FacultyImportLog.ImportStatus.SUCCESS,
            summary={
                'records_created': created_count,
                'records_updated': updated_count,
                'total_processed': len(self.valid_rows),
            },
            error_log=self.errors,
        )

        # Trigger automatic exam creation check
        self.academic_year.check_and_trigger_exam_creation()

        return import_log


# ═══════════════════════════════════════════════════════
# TEACHING ALLOCATION IMPORT
# ═══════════════════════════════════════════════════════

class TeachingAllocationImportService:
    """
    Parses a course allocation Excel/CSV file and creates
    FacultyTeachingAssignment records.

    The file has a multi-row header:
        Row 1: Sem | Class | Course Code | Course Title | (empty) | Faculty Allocation
        Row 2: (empty) | ... | ... | ... | Course Coordinator | Theory | Practical/Tutorial
        Row 3: (empty) | ... | ... | ... | (empty) | All | Batch-A | Batch-B | Batch-C

    Faculty are identified by alias (short form). The service resolves
    aliases by looking up FacultyMaster records:
        1. Match by short_form (primary)
        2. Fall back to name matching if short_form not found

    Theory faculty teach ALL students; Practical faculty are per-batch.
    """

    # Special values that indicate non-faculty entries
    SKIP_VALUES = {
        'external', 'online', 'class coordinator',
        'class\ncoordinator', 'tbd', 'tba', 'na', 'n/a', '-', ''
    }

    def __init__(self, file_obj, filename, academic_year):
        self.file_obj = file_obj
        self.filename = filename
        self.academic_year = academic_year
        self.errors = []
        self.warnings = []
        self.valid_rows = []
        self.file_format = 'xlsx' if filename.lower().endswith('.xlsx') else 'csv'

        # Build caches
        self._build_caches()

    def _build_caches(self):
        """Pre-load lookup caches for fast resolution."""
        # Faculty Master alias → FacultyMaster
        self.faculty_by_alias = {}
        self.faculty_by_name = {}
        for fm in FacultyMaster.objects.filter(academic_year=self.academic_year):
            if fm.short_form:
                self.faculty_by_alias[fm.short_form.upper()] = fm
            # Also index by name fragments for fallback
            name_key = fm.faculty_name.upper().strip()
            self.faculty_by_name[name_key] = fm

        # SemesterSubject by code for this AY
        self.subject_by_code = {}
        for ss in SemesterSubject.objects.filter(
            semester__academic_year=self.academic_year
        ).select_related('semester', 'subject'):
            self.subject_by_code[ss.subject_code.upper()] = ss

        # Semesters by number
        self.semesters = {}
        for sem in Semester.objects.filter(academic_year=self.academic_year):
            self.semesters[sem.number] = sem

    def validate(self):
        """Phase 1: Parse and validate without writing to DB."""
        self.errors = []
        self.warnings = []
        self.valid_rows = []

        try:
            if self.file_format == 'xlsx':
                self._parse_xlsx()
            else:
                self._parse_csv()
        except Exception as e:
            self.errors.append({
                'row': 0,
                'message': f"Failed to parse file: {str(e)}"
            })

        return self._get_summary()

    def _parse_xlsx(self):
        """Parse Excel course allocation file."""
        import openpyxl

        self.file_obj.seek(0)
        wb = openpyxl.load_workbook(self.file_obj, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.errors.append({'row': 0, 'message': "File is empty."})
            return

        # Find the header block
        data_start_idx = self._find_and_parse_header_xlsx(rows, start_from=0)
        if data_start_idx is None:
            self.errors.append({
                'row': 0,
                'message': (
                    "Could not find header row. Expected columns: "
                    "Sem, Class, Course Code, Course Title."
                )
            })
            return

        # Parse data rows
        current_semester = None

        for row_idx in range(data_start_idx, len(rows)):
            row = rows[row_idx]
            excel_row = row_idx + 1

            # Skip empty rows
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            # Check if this is a repeated header row (happens mid-file)
            if self._is_header_row(row):
                # Re-parse the header block (it may be 3 rows again)
                new_start = self._find_and_parse_header_xlsx(rows, start_from=row_idx)
                if new_start:
                    # Jump to the new data start; the for loop will continue from there
                    # We can't modify the loop variable, but we'll skip processed rows
                    data_start_idx = new_start
                continue

            # Skip rows that are part of a re-parsed header
            if row_idx < data_start_idx:
                continue

            self._process_data_row(row, excel_row, current_semester)

            # Update current semester from the row
            sem_val = self._get_cell(row, self.col_map.get('sem'))
            if sem_val:
                try:
                    current_semester = int(float(str(sem_val)))
                except (ValueError, TypeError):
                    pass

        wb.close()

    def _find_and_parse_header_xlsx(self, rows, start_from=0):
        """
        Find the multi-row header starting from a given index.
        Returns the index of the first data row, or None.

        Header structure:
            Row N:   Sem | Class | Course Code | Course Title | (empty) | Faculty Allocation
            Row N+1: ... | ... | ... | ... | Course Coordinator | Theory | Practical/Tutorial
            Row N+2: ... | ... | ... | ... | (empty) | All | Batch-A | Batch-B | Batch-C
        """
        self.col_map = getattr(self, 'col_map', {})

        for idx in range(start_from, min(start_from + 10, len(rows))):
            row = rows[idx]
            row_strs = [str(cell).strip().lower() if cell else '' for cell in row]

            if any(val in ('sem', 'semester') for val in row_strs):
                # Found main header row — map core columns
                for col_idx, val in enumerate(row_strs):
                    if val in ('sem', 'semester'):
                        self.col_map['sem'] = col_idx
                    elif val in ('class', 'class name', 'division'):
                        self.col_map['class'] = col_idx
                    elif val in ('course code', 'subject code', 'code'):
                        self.col_map['course_code'] = col_idx
                    elif val in ('course title', 'course name', 'subject name',
                                 'subject title', 'title'):
                        self.col_map['course_title'] = col_idx

                # Now look at sub-header rows to find faculty columns
                if idx + 1 < len(rows):
                    row2 = [str(c).strip().lower() if c else '' for c in rows[idx + 1]]
                    for col_idx, val in enumerate(row2):
                        if val in ('course coordinator', 'coordinator'):
                            self.col_map['coordinator'] = col_idx
                        elif val == 'theory':
                            self.col_map['theory'] = col_idx
                        elif val in ('practical/tutorial', 'practical', 'tutorial',
                                     'lab', 'practical / tutorial'):
                            self.col_map['practical_start'] = col_idx

                if idx + 2 < len(rows):
                    row3 = [str(c).strip().lower() if c else '' for c in rows[idx + 2]]
                    for col_idx, val in enumerate(row3):
                        if val == 'all':
                            self.col_map['theory_all'] = col_idx
                        elif val in ('batch-a', 'batch a', 'a'):
                            self.col_map['batch_a'] = col_idx
                        elif val in ('batch-b', 'batch b', 'b'):
                            self.col_map['batch_b'] = col_idx
                        elif val in ('batch-c', 'batch c', 'c'):
                            self.col_map['batch_c'] = col_idx

                # If we didn't find sub-header specific columns,
                # try to infer from positions
                if 'coordinator' not in self.col_map and 'course_title' in self.col_map:
                    self.col_map['coordinator'] = self.col_map['course_title'] + 1

                if 'theory_all' not in self.col_map and 'coordinator' in self.col_map:
                    self.col_map['theory_all'] = self.col_map['coordinator'] + 1

                if 'batch_a' not in self.col_map and 'theory_all' in self.col_map:
                    self.col_map['batch_a'] = self.col_map['theory_all'] + 1

                if 'batch_b' not in self.col_map and 'batch_a' in self.col_map:
                    self.col_map['batch_b'] = self.col_map['batch_a'] + 1

                if 'batch_c' not in self.col_map and 'batch_b' in self.col_map:
                    self.col_map['batch_c'] = self.col_map['batch_b'] + 1

                # Data starts after the header block (usually 3 rows)
                # Check for an empty separator row
                data_start = idx + 3
                if data_start < len(rows):
                    check_row = rows[data_start]
                    if not any(c is not None and str(c).strip() for c in check_row):
                        data_start += 1  # Skip empty separator

                return data_start

        return None

    def _is_header_row(self, row):
        """Check if a row is a repeated header row."""
        row_strs = [str(cell).strip().lower() if cell else '' for cell in row]
        return any(val in ('sem', 'semester') for val in row_strs) and \
               any(val in ('class', 'course code', 'course title') for val in row_strs)

    def _process_data_row(self, row, excel_row, fallback_semester):
        """Process a single data row, generating multiple assignment entries."""
        # Extract core fields
        sem_val = self._get_cell(row, self.col_map.get('sem'))
        class_val = self._get_cell(row, self.col_map.get('class'))
        code_val = self._get_cell(row, self.col_map.get('course_code'))
        title_val = self._get_cell(row, self.col_map.get('course_title'))
        coord_val = self._get_cell(row, self.col_map.get('coordinator'))
        theory_val = self._get_cell(row, self.col_map.get('theory_all'))
        batch_a_val = self._get_cell(row, self.col_map.get('batch_a'))
        batch_b_val = self._get_cell(row, self.col_map.get('batch_b'))
        batch_c_val = self._get_cell(row, self.col_map.get('batch_c'))

        # Determine semester
        semester_num = fallback_semester
        if sem_val:
            try:
                semester_num = int(float(str(sem_val)))
            except (ValueError, TypeError):
                pass

        # Skip rows without a course title
        if not title_val:
            return

        # Skip rows without a course code (elective placeholders)
        if not code_val:
            self.warnings.append({
                'row': excel_row,
                'message': f"No course code for '{title_val}'. Skipping."
            })
            return

        # Validate semester
        if semester_num is None:
            self.errors.append({
                'row': excel_row,
                'identifier': code_val,
                'message': "No semester number found."
            })
            return

        semester = self.semesters.get(semester_num)
        if not semester:
            self.errors.append({
                'row': excel_row,
                'identifier': code_val,
                'message': f"Semester {semester_num} not found in Academic Year {self.academic_year.name}."
            })
            return

        # Resolve subject
        code_upper = code_val.strip().upper()
        sem_subject = self.subject_by_code.get(code_upper)
        if not sem_subject:
            self.errors.append({
                'row': excel_row,
                'identifier': code_val,
                'message': (
                    f"Subject '{code_val}' not found in Semester {semester_num}. "
                    f"Import the Academic Structure first."
                )
            })
            return

        if not class_val:
            self.warnings.append({
                'row': excel_row,
                'identifier': code_val,
                'message': "No class name found. Using 'Unknown'."
            })
            classes = ['Unknown']
        else:
            import re
            classes = [c.strip() for c in re.split(r'[/+]', class_val) if c.strip()]

        # Build assignment entries for each faculty column
        assignments = []

        # Course Coordinator
        if coord_val:
            faculty, alias_raw = self._resolve_faculty(coord_val, excel_row)
            assignments.append({
                'teaching_type': FacultyTeachingAssignment.TeachingType.COORDINATOR,
                'faculty': faculty,
                'alias_raw': alias_raw,
                'is_coordinator': True,
            })

        # Theory (All students)
        if theory_val:
            faculty, alias_raw = self._resolve_faculty(theory_val, excel_row)
            assignments.append({
                'teaching_type': FacultyTeachingAssignment.TeachingType.THEORY,
                'faculty': faculty,
                'alias_raw': alias_raw,
                'is_coordinator': False,
            })

        # Practical/Tutorial Batch A
        if batch_a_val:
            faculty, alias_raw = self._resolve_faculty(batch_a_val, excel_row)
            assignments.append({
                'teaching_type': FacultyTeachingAssignment.TeachingType.PRACTICAL_BATCH_A,
                'faculty': faculty,
                'alias_raw': alias_raw,
                'is_coordinator': False,
            })

        # Practical/Tutorial Batch B
        if batch_b_val:
            faculty, alias_raw = self._resolve_faculty(batch_b_val, excel_row)
            assignments.append({
                'teaching_type': FacultyTeachingAssignment.TeachingType.PRACTICAL_BATCH_B,
                'faculty': faculty,
                'alias_raw': alias_raw,
                'is_coordinator': False,
            })

        # Practical/Tutorial Batch C
        if batch_c_val:
            faculty, alias_raw = self._resolve_faculty(batch_c_val, excel_row)
            assignments.append({
                'teaching_type': FacultyTeachingAssignment.TeachingType.PRACTICAL_BATCH_C,
                'faculty': faculty,
                'alias_raw': alias_raw,
                'is_coordinator': False,
            })

        for assign in assignments:
            for cls in classes:
                self.valid_rows.append({
                    'row': excel_row,
                    'semester_num': semester_num,
                    'semester_id': semester.id,
                    'semester_subject_id': sem_subject.id,
                    'course_code': code_upper,
                    'course_title': title_val,
                    'class_name': cls,
                    'teaching_type': assign['teaching_type'],
                    'faculty_id': assign['faculty'].id if assign['faculty'] else None,
                    'faculty_name': assign['faculty'].faculty_name if assign['faculty'] else '',
                    'faculty_email': assign['faculty'].email if assign['faculty'] else '',
                    'alias_raw': assign['alias_raw'],
                    'is_coordinator': assign['is_coordinator'],
                })

    def _resolve_faculty(self, alias_raw, excel_row):
        """
        Resolve a faculty alias to a FacultyMaster record.

        Strategy:
            1. Exact match on short_form (case-insensitive)
            2. Partial name match (fallback)
            3. Return None with warning if unresolvable

        Returns (FacultyMaster or None, raw_alias_string).
        """
        alias_clean = alias_raw.strip().replace('\n', ' ')

        # Skip special values
        if alias_clean.lower() in self.SKIP_VALUES:
            if alias_clean.lower() not in ('', '-'):
                self.warnings.append({
                    'row': excel_row,
                    'message': f"Skipping non-faculty value: '{alias_clean}'"
                })
            return None, alias_clean

        alias_upper = alias_clean.upper()

        # 1. Exact alias match
        faculty = self.faculty_by_alias.get(alias_upper)
        if faculty:
            return faculty, alias_clean

        # 2. Partial name match (check if alias appears in any faculty name)
        for name_key, fm in self.faculty_by_name.items():
            if alias_upper == name_key:
                return fm, alias_clean

        # Not found — warn but don't error (allow import to continue)
        self.warnings.append({
            'row': excel_row,
            'message': (
                f"Faculty alias '{alias_clean}' not found in Faculty Master. "
                f"Assignment will be created without faculty link."
            )
        })
        return None, alias_clean

    def _parse_csv(self):
        """Parse CSV course allocation file."""
        self.file_obj.seek(0)
        try:
            decoded = self.file_obj.read().decode('utf-8', errors='replace')
        except AttributeError:
            decoded = self.file_obj.read()

        io_string = io.StringIO(decoded)
        reader = csv.reader(io_string)
        all_rows = [tuple(row) for row in reader]

        if not all_rows:
            self.errors.append({'row': 0, 'message': "File is empty."})
            return

        data_start = self._find_and_parse_header_xlsx(all_rows, start_from=0)
        if data_start is None:
            self.errors.append({
                'row': 0,
                'message': "Could not find header row."
            })
            return

        current_semester = None
        for row_idx in range(data_start, len(all_rows)):
            row = all_rows[row_idx]
            excel_row = row_idx + 1

            if not any(str(c).strip() for c in row):
                continue

            if self._is_header_row(row):
                new_start = self._find_and_parse_header_xlsx(all_rows, start_from=row_idx)
                if new_start:
                    data_start = new_start
                continue

            if row_idx < data_start:
                continue

            sem_val = self._get_cell(row, self.col_map.get('sem'))
            if sem_val:
                try:
                    current_semester = int(float(str(sem_val)))
                except (ValueError, TypeError):
                    pass

            self._process_data_row(row, excel_row, current_semester)

    def _get_cell(self, row, col_idx):
        """Safely get a cell value."""
        if col_idx is None or col_idx >= len(row):
            return None
        val = row[col_idx]
        if val is None:
            return None
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() == 'none':
            return None
        return val_str

    def _get_summary(self):
        """Build validation summary."""
        # Count unique classes, subjects, faculty
        unique_classes = set()
        unique_subjects = set()
        unique_faculty = set()
        for r in self.valid_rows:
            unique_classes.add(r['class_name'])
            unique_subjects.add(r['course_code'])
            if r['faculty_id']:
                unique_faculty.add(r['faculty_id'])

        return {
            'total_valid': len(self.valid_rows),
            'total_errors': len(self.errors),
            'total_warnings': len(self.warnings),
            'unique_classes': len(unique_classes),
            'unique_subjects': len(unique_subjects),
            'unique_faculty': len(unique_faculty),
            'errors': self.errors[:100],
            'warnings': self.warnings[:100],
            'valid_sample': self.valid_rows[:20],
            'file_format': self.file_format,
        }

    @transaction.atomic
    def process_import(self, imported_by=None):
        """
        Phase 2: Commit validated assignments to the database.
        Uses get_or_create to prevent duplicates.
        """
        if not self.valid_rows:
            return FacultyImportLog.objects.create(
                academic_year=self.academic_year,
                import_type=FacultyImportLog.ImportType.TEACHING_ALLOCATION,
                original_filename=self.filename,
                file_format=self.file_format,
                imported_by=imported_by,
                status=FacultyImportLog.ImportStatus.FAILED,
                summary={'error': 'No valid rows to import.'},
                error_log=self.errors,
            )

        created_count = 0
        updated_count = 0

        for row_data in self.valid_rows:
            defaults = {
                'is_coordinator': row_data['is_coordinator'],
                'faculty_alias_raw': row_data['alias_raw'],
            }

            obj, created = FacultyTeachingAssignment.objects.update_or_create(
                academic_year=self.academic_year,
                semester_id=row_data['semester_id'],
                semester_subject_id=row_data['semester_subject_id'],
                class_name=row_data['class_name'],
                teaching_type=row_data['teaching_type'],
                faculty_id=row_data['faculty_id'],
                defaults=defaults,
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

        import_log = FacultyImportLog.objects.create(
            academic_year=self.academic_year,
            import_type=FacultyImportLog.ImportType.TEACHING_ALLOCATION,
            original_filename=self.filename,
            file_format=self.file_format,
            imported_by=imported_by,
            status=FacultyImportLog.ImportStatus.SUCCESS,
            summary={
                'assignments_created': created_count,
                'assignments_updated': updated_count,
                'total_processed': len(self.valid_rows),
            },
            error_log=self.errors,
        )

        # Trigger automatic exam creation check
        self.academic_year.check_and_trigger_exam_creation()

        return import_log
