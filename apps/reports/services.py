"""
Services for Reports Generation.
Supports CSV (text), PDF (reportlab), and Excel (openpyxl) output.
"""
import csv
import io
from django.db import transaction

from apps.reports.models import Report
from apps.core.models_audit import TextContent

from apps.seating.models import SeatingAllocation
from apps.duty_chart.models import DutyAssignment
from apps.marks.models import StudentMark
from apps.analysis.services import AnalysisService


class ReportGeneratorService:
    def __init__(self, exam, user):
        self.exam = exam
        self.user = user

    def _save_report(self, report_type, content_text, format_type='csv'):
        """Saves or updates a report and its TextContent."""
        with transaction.atomic():
            report = Report.objects.filter(exam=self.exam, report_type=report_type).first()
            if report:
                # Update existing text content
                report.content.content = content_text
                report.content.save()
                report.generated_by = self.user
                report.save()
            else:
                # Create new
                content = TextContent.objects.create(
                    title=f"{self.exam.name}_{report_type}",
                    content=content_text,
                    content_type=format_type,
                    module=f"reports_{report_type}",
                    related_object_id=str(self.exam.id),
                    related_model='Exam',
                    created_by=self.user,
                )
                report = Report.objects.create(
                    exam=self.exam,
                    report_type=report_type,
                    content=content,
                    generated_by=self.user
                )
            return report

    # ─── CSV Report Generators ───────────────────────────────────────

    def generate_seating_arrangement(self):
        """Generates a CSV text report of seating arrangements."""
        allocations = SeatingAllocation.objects.filter(plan__exam=self.exam).select_related(
            'schedule__subject', 'classroom', 'student'
        ).order_by('schedule__date', 'classroom__room_number', 'seat_number')
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Start Time', 'End Time', 'Subject', 'Classroom', 'Seat No', 'Student Roll No', 'Student Name'])
        
        for a in allocations:
            writer.writerow([
                a.schedule.date,
                a.schedule.start_time,
                a.schedule.end_time,
                a.schedule.subject.code,
                a.classroom.room_number,
                a.seat_number,
                a.student.roll_no,
                a.student.name
            ])
            
        return self._save_report('seating_arrangement', output.getvalue())

    def generate_duty_chart(self):
        """Generates a CSV text report of duty chart."""
        assignments = DutyAssignment.objects.filter(chart__exam=self.exam).select_related(
            'classroom', 'faculty'
        ).order_by('date', 'start_time', 'classroom__room_number')
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Date', 'Start Time', 'End Time', 'Classroom', 'Faculty Name', 'Faculty Role'])
        
        for a in assignments:
            writer.writerow([
                a.date,
                a.start_time,
                a.end_time,
                a.classroom.room_number,
                a.faculty.get_display_name(),
                a.faculty.get_role_display()
            ])
            
        return self._save_report('duty_chart', output.getvalue())

    def generate_marks_summary(self):
        """Generates a CSV text report of marks."""
        marks = StudentMark.objects.filter(task__exam=self.exam, task__status='locked').select_related(
            'student', 'task__subject'
        ).order_by('task__subject__code', 'student__roll_no')
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Subject Code', 'Student Roll No', 'Student Name', 'Absent', 'Total Marks', 'Component Details'])
        
        for m in marks:
            writer.writerow([
                m.task.subject.code,
                m.student.roll_no,
                m.student.name,
                'Yes' if m.is_absent else 'No',
                m.total_marks,
                str(m.component_marks)
            ])
            
        return self._save_report('marks_summary', output.getvalue())

    def generate_result_analysis(self):
        """Generates a CSV text report of analysis."""
        service = AnalysisService(self.exam)
        prog_data = service.get_program_analysis()
        subj_data = service.get_subject_analysis()
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['--- PROGRAM ANALYSIS ---'])
        writer.writerow(['Program', 'Average Marks', 'Pass Percentage'])
        for p in prog_data:
            writer.writerow([p['program'], p['average'], f"{p['pass_percentage']}%"])
            
        writer.writerow([])
        writer.writerow(['--- SUBJECT ANALYSIS ---'])
        writer.writerow(['Subject', 'Average Marks'])
        for s in subj_data:
            writer.writerow([s['subject'], s['average']])
            
        return self._save_report('result_analysis', output.getvalue())

    # ─── PDF Report Generators ───────────────────────────────────────

    def _get_pdf_buffer(self, title, headers, rows):
        """Helper to generate a PDF table from headers and rows using reportlab."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title=title)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        elements.append(Paragraph(title, styles['Title']))
        elements.append(Spacer(1, 0.3 * inch))

        # Table
        table_data = [headers] + rows
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0d9488')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return buf

    def generate_seating_pdf(self):
        """Generate seating arrangement as a PDF."""
        allocations = SeatingAllocation.objects.filter(plan__exam=self.exam).select_related(
            'schedule__subject', 'classroom', 'student'
        ).order_by('schedule__date', 'classroom__room_number', 'seat_number')

        headers = ['Date', 'Start', 'End', 'Subject', 'Room', 'Seat', 'Roll No', 'Name']
        rows = [[
            str(a.schedule.date), str(a.schedule.start_time), str(a.schedule.end_time),
            a.schedule.subject.code, a.classroom.room_number, a.seat_number,
            a.student.roll_no, a.student.name
        ] for a in allocations]

        return self._get_pdf_buffer(f'Seating Arrangement — {self.exam.name}', headers, rows)

    def generate_duty_chart_pdf(self):
        """Generate duty chart as a PDF."""
        assignments = DutyAssignment.objects.filter(chart__exam=self.exam).select_related(
            'classroom', 'faculty'
        ).order_by('date', 'start_time', 'classroom__room_number')

        headers = ['Date', 'Start', 'End', 'Room', 'Faculty', 'Role']
        rows = [[
            str(a.date), str(a.start_time), str(a.end_time),
            a.classroom.room_number, a.faculty.get_display_name(), a.faculty.get_role_display()
        ] for a in assignments]

        return self._get_pdf_buffer(f'Duty Chart — {self.exam.name}', headers, rows)

    def generate_marks_pdf(self):
        """Generate marks summary as a PDF."""
        marks = StudentMark.objects.filter(task__exam=self.exam, task__status='locked').select_related(
            'student', 'task__subject'
        ).order_by('task__subject__code', 'student__roll_no')

        headers = ['Subject', 'Roll No', 'Name', 'Absent', 'Total Marks']
        rows = [[
            m.task.subject.code, m.student.roll_no, m.student.name,
            'Yes' if m.is_absent else 'No', str(m.total_marks)
        ] for m in marks]

        return self._get_pdf_buffer(f'Marks Summary — {self.exam.name}', headers, rows)

    def generate_analysis_pdf(self):
        """Generate analysis report as a PDF."""
        service = AnalysisService(self.exam)
        prog_data = service.get_program_analysis()

        headers = ['Program', 'Average Marks', 'Pass %', 'Students']
        rows = [[
            p['program'], str(p['average']), f"{p['pass_percentage']}%", str(p['student_count'])
        ] for p in prog_data]

        return self._get_pdf_buffer(f'Result Analysis — {self.exam.name}', headers, rows)

    # ─── Excel Report Generators ─────────────────────────────────────

    def _get_excel_buffer(self, title, headers, rows):
        """Helper to generate an Excel workbook from headers and rows using openpyxl."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel limits sheet name to 31 chars

        # Header styling
        header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in rows:
                if col_idx - 1 < len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1])))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def generate_seating_excel(self):
        """Generate seating arrangement as an Excel file."""
        allocations = SeatingAllocation.objects.filter(plan__exam=self.exam).select_related(
            'schedule__subject', 'classroom', 'student'
        ).order_by('schedule__date', 'classroom__room_number', 'seat_number')

        headers = ['Date', 'Start Time', 'End Time', 'Subject', 'Classroom', 'Seat No', 'Roll No', 'Name']
        rows = [[
            str(a.schedule.date), str(a.schedule.start_time), str(a.schedule.end_time),
            a.schedule.subject.code, a.classroom.room_number, a.seat_number,
            a.student.roll_no, a.student.name
        ] for a in allocations]

        return self._get_excel_buffer('Seating Arrangement', headers, rows)

    def generate_duty_chart_excel(self):
        """Generate duty chart as an Excel file."""
        assignments = DutyAssignment.objects.filter(chart__exam=self.exam).select_related(
            'classroom', 'faculty'
        ).order_by('date', 'start_time', 'classroom__room_number')

        headers = ['Date', 'Start Time', 'End Time', 'Classroom', 'Faculty Name', 'Role']
        rows = [[
            str(a.date), str(a.start_time), str(a.end_time),
            a.classroom.room_number, a.faculty.get_display_name(), a.faculty.get_role_display()
        ] for a in assignments]

        return self._get_excel_buffer('Duty Chart', headers, rows)

    def generate_marks_excel(self):
        """Generate marks summary as an Excel file."""
        marks = StudentMark.objects.filter(task__exam=self.exam, task__status='locked').select_related(
            'student', 'task__subject'
        ).order_by('task__subject__code', 'student__roll_no')

        headers = ['Subject Code', 'Roll No', 'Student Name', 'Absent', 'Total Marks']
        rows = [[
            m.task.subject.code, m.student.roll_no, m.student.name,
            'Yes' if m.is_absent else 'No', float(m.total_marks)
        ] for m in marks]

        return self._get_excel_buffer('Marks Summary', headers, rows)

    def generate_analysis_excel(self):
        """Generate analysis report as an Excel file."""
        service = AnalysisService(self.exam)
        prog_data = service.get_program_analysis()

        headers = ['Program', 'Average Marks', 'Pass %', 'Student Count']
        rows = [[
            p['program'], p['average'], p['pass_percentage'], p['student_count']
        ] for p in prog_data]

        return self._get_excel_buffer('Result Analysis', headers, rows)
